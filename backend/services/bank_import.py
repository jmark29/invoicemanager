"""Bank statement XLSX import service.

Parses bank statement exports with 7 columns:
  Buchungstag, Wertstellung, Umsatzart, Buchungstext, Betrag, RK, Buchungsjahr

Auto-matches transactions to cost categories by searching Buchungstext for
bank_keywords (case-insensitive). Extracts invoice references via regex
(e.g., "ZAHLUNGSGRUND: INV320", "INVOICE  AEO000811").
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime

logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.models.bank_transaction import BankTransaction
from backend.models.cost_category import CostCategory

# Invoice reference patterns found in Buchungstext
INVOICE_REF_PATTERNS = [
    re.compile(r"ZAHLUNGSGRUND:\s*(INV\d+)", re.IGNORECASE),
    re.compile(r"INVOICE\s+(AEO\d+)", re.IGNORECASE),
    re.compile(r"INVOICE\s+(INV\d+)", re.IGNORECASE),
    re.compile(r"RE\.?\s*NR\.?\s*:?\s*(\d+/\d+)", re.IGNORECASE),  # e.g. "01/2025"
]


@dataclass
class ParsedBankTransaction:
    """A single parsed bank transaction."""

    booking_date: date
    value_date: date | None
    transaction_type: str | None
    description: str
    amount_eur: float  # negative for outgoing
    booking_year: int | None
    matched_category_id: str | None = None
    extracted_reference: str | None = None


@dataclass
class PotentialDuplicate:
    """A transaction that was identified as a potential duplicate."""

    booking_date: date
    amount_eur: float
    description: str


@dataclass
class BankImportResult:
    """Result of a bank statement XLSX import operation."""

    imported: int = 0
    skipped_duplicate: int = 0
    auto_matched: int = 0  # category matches
    invoice_auto_matched: int = 0  # invoice auto-links (high confidence)
    invoice_suggested: int = 0  # invoice suggestions (medium confidence)
    transactions: list[ParsedBankTransaction] = field(default_factory=list)
    potential_duplicates: list[PotentialDuplicate] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _parse_german_date(value) -> date | None:
    """Parse a date from bank XLSX (may be datetime, date, or DD.MM.YYYY string)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_amount(value) -> float | None:
    """Parse a numeric amount, handling German number format (1.234,56)."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # German format: "1.234,56" or "-1.234,56"
        cleaned = value.strip().replace(".", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _extract_invoice_reference(description: str) -> str | None:
    """Extract an invoice reference from the bank description text."""
    for pattern in INVOICE_REF_PATTERNS:
        m = pattern.search(description)
        if m:
            return m.group(1)
    return None


def _match_category(
    description: str,
    categories: list[CostCategory],
) -> str | None:
    """Match a bank transaction to a cost category by searching for keywords."""
    desc_lower = description.lower()
    for cat in categories:
        for keyword in cat.bank_keywords:
            if keyword.lower() in desc_lower:
                return cat.id
    return None


def parse_bank_xlsx(file_path: str) -> BankImportResult:
    """Parse a bank statement XLSX file and return structured transaction data.

    Does NOT write to the database — use ``import_bank_transactions`` for that.
    """
    result = BankImportResult()

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        result.errors.append("No active sheet found in workbook")
        wb.close()
        return result

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue

        buchungstag, wertstellung, umsatzart, buchungstext, betrag = row[:5]
        rk = row[5] if len(row) > 5 else None
        buchungsjahr = row[6] if len(row) > 6 else None

        # Skip empty rows
        if not buchungstag and not buchungstext:
            continue

        # Parse booking date
        booking_date = _parse_german_date(buchungstag)
        if not booking_date:
            result.errors.append(f"Invalid booking date: {buchungstag}")
            continue

        # Parse value date
        value_date = _parse_german_date(wertstellung)

        # Parse amount
        amount = _parse_amount(betrag)
        if amount is None:
            result.errors.append(f"Invalid amount on {buchungstag}: {betrag}")
            continue

        description = str(buchungstext) if buchungstext else ""
        reference = _extract_invoice_reference(description)

        booking_year = None
        if buchungsjahr is not None:
            try:
                booking_year = int(buchungsjahr)
            except (ValueError, TypeError):
                pass

        tx = ParsedBankTransaction(
            booking_date=booking_date,
            value_date=value_date,
            transaction_type=str(umsatzart) if umsatzart else None,
            description=description,
            amount_eur=amount,
            booking_year=booking_year,
            extracted_reference=reference,
        )
        result.transactions.append(tx)

    wb.close()
    return result


def import_bank_transactions(
    file_path: str,
    db: Session,
    force_import_all: bool = False,
) -> BankImportResult:
    """Parse bank statement XLSX and import transactions into the database.

    Auto-matches categories by bank_keywords. Skips duplicates based on
    (booking_date, amount_eur, description) combination unless
    ``force_import_all`` is True, in which case duplicates are imported anyway.
    """
    result = parse_bank_xlsx(file_path)

    # Load active categories with bank_keywords for matching
    categories = (
        db.query(CostCategory)
        .filter(CostCategory.active.is_(True))
        .all()
    )
    categories_with_keywords = [c for c in categories if c.bank_keywords]

    # Build set of existing transactions for duplicate detection
    existing = db.query(
        BankTransaction.booking_date,
        BankTransaction.amount_eur,
        BankTransaction.description,
    ).all()
    existing_set = {(row[0], row[1], row[2]) for row in existing}

    imported = []
    for tx in result.transactions:
        # Duplicate check
        key = (tx.booking_date, tx.amount_eur, tx.description)
        if key in existing_set and not force_import_all:
            result.skipped_duplicate += 1
            result.potential_duplicates.append(
                PotentialDuplicate(
                    booking_date=tx.booking_date,
                    amount_eur=tx.amount_eur,
                    description=tx.description,
                )
            )
            continue

        # Auto-match category
        matched_cat_id = _match_category(tx.description, categories_with_keywords)
        tx.matched_category_id = matched_cat_id
        if matched_cat_id:
            result.auto_matched += 1

        record = BankTransaction(
            booking_date=tx.booking_date,
            value_date=tx.value_date,
            transaction_type=tx.transaction_type,
            description=tx.description,
            amount_eur=tx.amount_eur,
            reference=tx.extracted_reference,
            category_id=matched_cat_id,
        )
        db.add(record)
        imported.append(record)
        existing_set.add(key)  # Prevent duplicates within same import

    if imported:
        db.commit()

    result.imported = len(imported)

    # Auto-match imported transactions to existing provider invoices
    if imported:
        from backend.services.transaction_matching import auto_match_after_bank_import
        imported_ids = [r.id for r in imported]
        match_stats = auto_match_after_bank_import(imported_ids, db)
        result.invoice_auto_matched = match_stats.auto_matched
        result.invoice_suggested = match_stats.suggested

    logger.info(
        "Bank import: %d imported, %d duplicates skipped, %d category-matched",
        result.imported, result.skipped_duplicate, result.auto_matched,
    )
    return result
