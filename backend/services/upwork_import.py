"""Upwork XLSX import service.

Parses the Upwork transaction export (sheet "data", 9 columns) and imports
transactions into the database. Month assignment is determined by the period
END date extracted from the transaction summary field.

Columns: Date, Transaction ID, Transaction type, Transaction summary details,
         Description 1, Ref ID, Amount in local currency, Currency, Payment method
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime

logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.models.upwork_transaction import UpworkTransaction

# Month abbreviation -> number
MONTH_MAP = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

# Pattern 1: "Invoice for Feb 16-Feb 22, 2026" or "Invoice for Feb 16, Feb 22, 2026"
PERIOD_RE = re.compile(
    r"Invoice for (\w+ \d+)[,\-]\s*(\w+ \d+),?\s*(\d{4})"
)

# Pattern 2: "Invoice for Dec 29, 2025-Jan 4, 2026" (cross-year)
PERIOD_CROSS_YEAR_RE = re.compile(
    r"Invoice for (\w+ \d+),?\s*(\d{4})\s*-\s*(\w+ \d+),?\s*(\d{4})"
)


@dataclass
class ImportedTransaction:
    """A single parsed Upwork transaction."""

    tx_id: str
    tx_date: date
    tx_type: str | None
    description: str | None
    period_start: date | None
    period_end: date | None
    amount_eur: float
    assigned_month: str | None  # "YYYY-MM" from period end date
    freelancer_name: str | None
    contract_ref: str | None


@dataclass
class UpworkImportResult:
    """Result of an Upwork XLSX import operation."""

    imported: int = 0
    skipped_duplicate: int = 0
    skipped_no_amount: int = 0
    skipped_no_period: int = 0
    transactions: list[ImportedTransaction] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _parse_date(value) -> date | None:
    """Parse a date value from the XLSX cell."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%b %d, %Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_period(summary: str) -> tuple[date | None, date | None]:
    """Extract period start and end dates from Upwork transaction summary.

    Returns (period_start, period_end) or (None, None) if unparseable.
    """
    if not summary:
        return None, None

    # Try cross-year pattern first (more specific)
    m = PERIOD_CROSS_YEAR_RE.search(summary)
    if m:
        start_parts = m.group(1).split()
        start_year = int(m.group(2))
        end_parts = m.group(3).split()
        end_year = int(m.group(4))

        start_month = MONTH_MAP.get(start_parts[0], 0)
        start_day = int(start_parts[1])
        end_month = MONTH_MAP.get(end_parts[0], 0)
        end_day = int(end_parts[1])

        if start_month and end_month:
            return (
                date(start_year, start_month, start_day),
                date(end_year, end_month, end_day),
            )
        return None, None

    # Try standard pattern
    m = PERIOD_RE.search(summary)
    if m:
        start_parts = m.group(1).split()
        end_parts = m.group(2).split()
        year = int(m.group(3))

        start_month = MONTH_MAP.get(start_parts[0], 0)
        start_day = int(start_parts[1])
        end_month = MONTH_MAP.get(end_parts[0], 0)
        end_day = int(end_parts[1])

        if start_month and end_month:
            return (
                date(year, start_month, start_day),
                date(year, end_month, end_day),
            )

    return None, None


def parse_upwork_xlsx(file_path: str) -> UpworkImportResult:
    """Parse an Upwork XLSX file and return structured transaction data.

    Does NOT write to the database — use ``import_upwork_transactions`` for that.
    """
    result = UpworkImportResult()

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if "data" not in wb.sheetnames:
        result.errors.append(f"Sheet 'data' not found. Available: {wb.sheetnames}")
        wb.close()
        return result

    ws = wb["data"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 7:
            continue

        tx_date_raw, tx_id_raw, tx_type, tx_summary, tx_desc, ref_id, amount_raw = row[:7]

        # Skip rows without transaction ID or amount
        if not tx_id_raw or amount_raw is None:
            result.skipped_no_amount += 1
            continue

        # Normalize tx_id
        tx_id = str(int(tx_id_raw)) if isinstance(tx_id_raw, (int, float)) else str(tx_id_raw).strip()

        # Parse amount (EUR despite column name "Amount in local currency")
        try:
            amount = float(amount_raw)
        except (ValueError, TypeError):
            result.errors.append(f"Invalid amount for tx {tx_id}: {amount_raw}")
            continue

        # Parse dates
        tx_date = _parse_date(tx_date_raw)
        if not tx_date:
            result.errors.append(f"Invalid date for tx {tx_id}: {tx_date_raw}")
            continue

        # Parse period from summary
        period_start, period_end = _parse_period(str(tx_summary) if tx_summary else "")

        # Determine assigned month from period end date
        assigned_month = None
        if period_end:
            assigned_month = f"{period_end.year}-{period_end.month:02d}"
        else:
            result.skipped_no_period += 1

        tx = ImportedTransaction(
            tx_id=tx_id,
            tx_date=tx_date,
            tx_type=str(tx_type) if tx_type else None,
            description=str(tx_desc) if tx_desc else None,
            period_start=period_start,
            period_end=period_end,
            amount_eur=amount,
            assigned_month=assigned_month,
            freelancer_name=None,  # Could be extracted from description
            contract_ref=str(ref_id) if ref_id else None,
        )
        result.transactions.append(tx)

    wb.close()
    return result


def import_upwork_transactions(
    file_path: str,
    db: Session,
    category_id: str | None = None,
) -> UpworkImportResult:
    """Parse Upwork XLSX and import transactions into the database.

    Skips duplicates based on tx_id. Optionally assigns a category_id to
    all imported transactions.
    """
    result = parse_upwork_xlsx(file_path)

    # Check for existing tx_ids in a single query
    existing_ids = set()
    if result.transactions:
        all_ids = [tx.tx_id for tx in result.transactions]
        existing = (
            db.query(UpworkTransaction.tx_id)
            .filter(UpworkTransaction.tx_id.in_(all_ids))
            .all()
        )
        existing_ids = {row[0] for row in existing}

    imported = []
    for tx in result.transactions:
        if tx.tx_id in existing_ids:
            result.skipped_duplicate += 1
            continue

        record = UpworkTransaction(
            tx_id=tx.tx_id,
            tx_date=tx.tx_date,
            tx_type=tx.tx_type,
            description=tx.description,
            period_start=tx.period_start,
            period_end=tx.period_end,
            amount_eur=tx.amount_eur,
            assigned_month=tx.assigned_month,
            contract_ref=tx.contract_ref,
            freelancer_name=tx.freelancer_name,
            category_id=category_id,
        )
        db.add(record)
        imported.append(record)

    if imported:
        db.commit()

    result.imported = len(imported)
    logger.info(
        "Upwork import: %d imported, %d duplicates skipped",
        result.imported, result.skipped_duplicate,
    )
    return result
