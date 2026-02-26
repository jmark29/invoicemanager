"""Tests for Phase 3A — Upwork and bank import services."""

import os
import tempfile
from datetime import date

import pytest
from openpyxl import Workbook
from sqlalchemy.orm import Session

from backend.models.bank_transaction import BankTransaction
from backend.models.cost_category import CostCategory
from backend.models.upwork_transaction import UpworkTransaction
from backend.services.bank_import import (
    _extract_invoice_reference,
    _match_category,
    _parse_amount,
    _parse_german_date,
    import_bank_transactions,
    parse_bank_xlsx,
)
from backend.services.upwork_import import (
    _parse_period,
    import_upwork_transactions,
    parse_upwork_xlsx,
)


# ─── Upwork Period Parsing ──────────────────────────────────────


class TestUpworkPeriodParsing:
    def test_standard_period(self):
        start, end = _parse_period("Invoice for Feb 16-Feb 22, 2026")
        assert start == date(2026, 2, 16)
        assert end == date(2026, 2, 22)

    def test_standard_period_with_comma(self):
        start, end = _parse_period("Invoice for Feb 16, Feb 22, 2026")
        assert start == date(2026, 2, 16)
        assert end == date(2026, 2, 22)

    def test_cross_year_period(self):
        start, end = _parse_period("Invoice for Dec 29, 2025-Jan 4, 2026")
        assert start == date(2025, 12, 29)
        assert end == date(2026, 1, 4)

    def test_cross_month_period(self):
        start, end = _parse_period("Invoice for Jan 27-Feb 2, 2026")
        assert start == date(2026, 1, 27)
        assert end == date(2026, 2, 2)

    def test_empty_summary(self):
        start, end = _parse_period("")
        assert start is None
        assert end is None

    def test_no_match(self):
        start, end = _parse_period("Some random text without period")
        assert start is None
        assert end is None

    def test_month_assignment_by_end_date(self):
        """Month assignment should use the period END date."""
        _, end = _parse_period("Invoice for Dec 29, 2025-Jan 4, 2026")
        assert end is not None
        assigned = f"{end.year}-{end.month:02d}"
        assert assigned == "2026-01"


# ─── Upwork XLSX Parsing ────────────────────────────────────────


def _create_upwork_xlsx(rows: list[list], path: str) -> None:
    """Helper to create a test Upwork XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    # Header
    ws.append([
        "Date", "Transaction ID", "Transaction type",
        "Transaction summary details", "Description 1", "Ref ID",
        "Amount in local currency", "Currency", "Payment method",
    ])
    for row in rows:
        ws.append(row)
    wb.save(path)


class TestUpworkXlsxParsing:
    def test_parse_basic_transactions(self, tmp_path):
        xlsx_path = str(tmp_path / "upwork.xlsx")
        _create_upwork_xlsx([
            [date(2026, 2, 20), 12345, "Hourly",
             "Invoice for Feb 16-Feb 22, 2026", "Mobile Dev", "REF1",
             150.50, "EUR", "Direct"],
            [date(2026, 2, 27), 12346, "Hourly",
             "Invoice for Feb 23-Mar 1, 2026", "Mobile Dev", "REF2",
             200.00, "EUR", "Direct"],
        ], xlsx_path)

        result = parse_upwork_xlsx(xlsx_path)
        assert len(result.transactions) == 2
        assert result.transactions[0].tx_id == "12345"
        assert result.transactions[0].amount_eur == 150.50
        assert result.transactions[0].assigned_month == "2026-02"
        assert result.transactions[1].assigned_month == "2026-03"

    def test_skip_no_amount(self, tmp_path):
        xlsx_path = str(tmp_path / "upwork.xlsx")
        _create_upwork_xlsx([
            [date(2026, 2, 20), 12345, "Hourly",
             "Invoice for Feb 16-Feb 22, 2026", "Dev", "REF1",
             None, "EUR", "Direct"],
        ], xlsx_path)

        result = parse_upwork_xlsx(xlsx_path)
        assert len(result.transactions) == 0
        assert result.skipped_no_amount == 1

    def test_wrong_sheet_name(self, tmp_path):
        xlsx_path = str(tmp_path / "upwork.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"  # Wrong name
        ws.append(["Date", "Transaction ID"])
        wb.save(xlsx_path)

        result = parse_upwork_xlsx(xlsx_path)
        assert len(result.errors) == 1
        assert "Sheet 'data' not found" in result.errors[0]


class TestUpworkImport:
    def test_import_and_deduplicate(self, db_session: Session, tmp_path):
        xlsx_path = str(tmp_path / "upwork.xlsx")
        _create_upwork_xlsx([
            [date(2026, 2, 20), 12345, "Hourly",
             "Invoice for Feb 16-Feb 22, 2026", "Dev", "REF1",
             150.50, "EUR", "Direct"],
        ], xlsx_path)

        # First import
        result1 = import_upwork_transactions(xlsx_path, db_session)
        assert result1.imported == 1
        assert result1.skipped_duplicate == 0

        # Second import — should skip duplicate
        result2 = import_upwork_transactions(xlsx_path, db_session)
        assert result2.imported == 0
        assert result2.skipped_duplicate == 1

        # Verify DB
        txs = db_session.query(UpworkTransaction).all()
        assert len(txs) == 1
        assert txs[0].tx_id == "12345"
        assert txs[0].amount_eur == 150.50

    def test_import_with_category(self, db_session: Session, tmp_path):
        # Create a category first
        cat = CostCategory(
            id="upwork",
            name="Mobile Dev",
            billing_cycle="weekly",
            cost_type="upwork",
        )
        db_session.add(cat)
        db_session.commit()

        xlsx_path = str(tmp_path / "upwork.xlsx")
        _create_upwork_xlsx([
            [date(2026, 2, 20), 99999, "Hourly",
             "Invoice for Feb 16-Feb 22, 2026", "Dev", "REF1",
             100.00, "EUR", "Direct"],
        ], xlsx_path)

        result = import_upwork_transactions(xlsx_path, db_session, category_id="upwork")
        assert result.imported == 1
        tx = db_session.query(UpworkTransaction).first()
        assert tx.category_id == "upwork"


# ─── Bank Import — Helper Functions ─────────────────────────────


class TestBankHelpers:
    def test_parse_german_date_string(self):
        assert _parse_german_date("15.03.2025") == date(2025, 3, 15)

    def test_parse_german_date_iso(self):
        assert _parse_german_date("2025-03-15") == date(2025, 3, 15)

    def test_parse_german_date_datetime(self):
        from datetime import datetime
        assert _parse_german_date(datetime(2025, 3, 15)) == date(2025, 3, 15)

    def test_parse_german_date_invalid(self):
        assert _parse_german_date("not a date") is None

    def test_parse_amount_float(self):
        assert _parse_amount(1234.56) == 1234.56

    def test_parse_amount_german_string(self):
        assert _parse_amount("1.234,56") == 1234.56

    def test_parse_amount_negative_german(self):
        assert _parse_amount("-8.295,00") == -8295.0

    def test_parse_amount_invalid(self):
        assert _parse_amount("abc") is None

    def test_extract_zahlungsgrund(self):
        ref = _extract_invoice_reference("ZAHLUNGSGRUND: INV320 some text")
        assert ref == "INV320"

    def test_extract_invoice_aeologic(self):
        ref = _extract_invoice_reference("INVOICE  AEO000811 payment")
        assert ref == "AEO000811"

    def test_extract_no_match(self):
        ref = _extract_invoice_reference("Regular transfer description")
        assert ref is None


class TestBankCategoryMatching:
    def test_match_by_keyword(self):
        cat = CostCategory(
            id="kaletsch",
            name="Cloud Engineer",
            billing_cycle="quarterly",
            cost_type="distributed",
        )
        cat.bank_keywords = ["kaletsch", "cloud engineer"]

        result = _match_category("Transfer to KALETSCH COMPANY", [cat])
        assert result == "kaletsch"

    def test_no_match(self):
        cat = CostCategory(
            id="kaletsch",
            name="Cloud Engineer",
            billing_cycle="quarterly",
            cost_type="distributed",
        )
        cat.bank_keywords = ["kaletsch"]

        result = _match_category("Transfer to someone else", [cat])
        assert result is None

    def test_case_insensitive(self):
        cat = CostCategory(
            id="aeologic",
            name="Aeologic",
            billing_cycle="irregular",
            cost_type="direct",
        )
        cat.bank_keywords = ["aeologic"]

        result = _match_category("INVOICE  AEO000811 AEOLOGIC TECHNOLOGIES", [cat])
        assert result == "aeologic"


# ─── Bank XLSX Parsing ──────────────────────────────────────────


def _create_bank_xlsx(rows: list[list], path: str) -> None:
    """Helper to create a test bank statement XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Buchungstag", "Wertstellung", "Umsatzart", "Buchungstext",
               "Betrag", "RK", "Buchungsjahr"])
    for row in rows:
        ws.append(row)
    wb.save(path)


class TestBankXlsxParsing:
    def test_parse_basic_transactions(self, tmp_path):
        xlsx_path = str(tmp_path / "bank.xlsx")
        _create_bank_xlsx([
            [date(2025, 1, 6), date(2025, 1, 6), "Überweisung",
             "ZAHLUNGSGRUND: INV307 KALETSCH COMPANY", -8295.00, None, 2025],
            [date(2025, 3, 31), date(2025, 3, 31), "Überweisung",
             "INVOICE  AEO000741 AEOLOGIC TECHNOLOGIES", -1036.28, None, 2025],
        ], xlsx_path)

        result = parse_bank_xlsx(xlsx_path)
        assert len(result.transactions) == 2
        assert result.transactions[0].amount_eur == -8295.00
        assert result.transactions[0].extracted_reference == "INV307"
        assert result.transactions[1].extracted_reference == "AEO000741"

    def test_skip_empty_rows(self, tmp_path):
        xlsx_path = str(tmp_path / "bank.xlsx")
        _create_bank_xlsx([
            [None, None, None, None, None, None, None],
            [date(2025, 1, 6), None, "Überweisung", "Test", -100.00, None, 2025],
        ], xlsx_path)

        result = parse_bank_xlsx(xlsx_path)
        assert len(result.transactions) == 1


class TestBankImport:
    def test_import_and_deduplicate(self, db_session: Session, tmp_path):
        xlsx_path = str(tmp_path / "bank.xlsx")
        _create_bank_xlsx([
            [date(2025, 1, 6), date(2025, 1, 6), "Überweisung",
             "KALETSCH COMPANY payment", -8295.00, None, 2025],
        ], xlsx_path)

        result1 = import_bank_transactions(xlsx_path, db_session)
        assert result1.imported == 1

        result2 = import_bank_transactions(xlsx_path, db_session)
        assert result2.imported == 0
        assert result2.skipped_duplicate == 1

    def test_auto_match_category(self, db_session: Session, tmp_path):
        # Create category with keywords
        cat = CostCategory(
            id="kaletsch",
            name="Cloud Engineer",
            billing_cycle="quarterly",
            cost_type="distributed",
        )
        cat.bank_keywords = ["kaletsch"]
        db_session.add(cat)
        db_session.commit()

        xlsx_path = str(tmp_path / "bank.xlsx")
        _create_bank_xlsx([
            [date(2025, 1, 6), date(2025, 1, 6), "Überweisung",
             "KALETSCH COMPANY INV307", -8295.00, None, 2025],
        ], xlsx_path)

        result = import_bank_transactions(xlsx_path, db_session)
        assert result.imported == 1
        assert result.auto_matched == 1

        tx = db_session.query(BankTransaction).first()
        assert tx.category_id == "kaletsch"
